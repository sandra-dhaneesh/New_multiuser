from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render,redirect
from django.contrib.auth.models import User, auth
from django.utils.text import capfirst
from django.contrib import messages
from . models import *
import json
from django.core.exceptions import ObjectDoesNotExist
from django.http.response import JsonResponse
from django.utils.crypto import get_random_string
from datetime import date
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.template.response import TemplateResponse
from openpyxl import load_workbook
from openpyxl import Workbook
from django.http import Http404
from django.http.response import JsonResponse, HttpResponse
from xhtml2pdf import pisa

from django.template.loader import get_template
from django.db.models import F

from django.http import JsonResponse
from django.db.models import Sum
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from io import BytesIO
from reportlab.pdfgen import canvas

from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import get_object_or_404
from django.db import transaction
from django.core.mail import send_mail, EmailMultiAlternatives
from django.template.loader import render_to_string
from django.utils.html import strip_tags
from .models import company

from django.db.models import Max

# Create your views here.

def generate_pdf_content():
    # Replace this with your actual PDF generation logic
    buffer = BytesIO()
    p = canvas.Canvas(buffer)
    p.drawString(100, 100, "This is a sample PDF content.")
    p.showPage()
    p.save()
    buffer.seek(0)
    return buffer

def home(request):
  return render(request, 'home.html')


def about(request):
  return render(request, 'about.html')

def contact(request):
  return render(request, 'contact.html')


def service(request):
  return render(request, 'service.html')

def parties_default(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  return render(request, 'parties_default.html',{'staff':staff, 'cmp':cmp})


def homepage(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  context={
   'staff':staff,
    'cmp':cmp,
  }
  return render(request, 'companyhome.html',context)


def staffhome(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  context={
   'staff':staff,
    'cmp':cmp,
  }
  return render(request, 'staffhome.html',context)

def register(request):
  return render(request, 'register.html')

def registercompany(request):
  return render(request, 'registercompany.html')

def registerstaff(request):
  return render(request, 'registerstaff.html')

def login(request):
  return render(request, 'login.html')

def logout(request):
    auth.logout(request)
    return redirect('/')


def registeruser(request):
    if request.method == 'POST':
        first_name = request.POST['fname']
        last_name = request.POST['lname']
        user_name = request.POST['username']
        email_id = request.POST['email']
        mobile = request.POST['phoneno']
        passw = request.POST['pass']
        c_passw = request.POST['re_pass']
        profile_pic=request.FILES.get('image')

        if passw != c_passw:
            messages.error(request, 'Passwords do not match')
            return redirect('register')

        if User.objects.filter(username=user_name).exists():
            messages.error(request, 'Sorry, Username already exists')
            return redirect('register')

        if User.objects.filter(email=email_id).exists():
            messages.error(request, 'Sorry, Email already exists')
            return redirect('register')

        # If everything is okay, save the data
        user_data = User.objects.create_user(
            first_name=first_name,
            last_name=last_name,
            username=user_name,
            email=email_id,
            password=passw
        )
        user_data.save()

        data = User.objects.get(id=user_data.id)
        cust_data = company(contact=mobile,profile_pic=profile_pic, user=data)
        cust_data.save()

        demo_staff=staff_details(company=cust_data,
                                   email=email_id,
                                   position='company',
                                   user_name=user_name,
                                   password=passw,
                                   contact=mobile)
        demo_staff.save()

        # messages.success(request, 'Registration successful')
        return redirect('registercompany')

    return render(request, 'register.html')
  
  
def add_company(request):
  
  if request.method == 'POST':
    email=request.POST['email']
    user=User.objects.get(email=email)
    
    c =company.objects.get(user = user)
    c.company_name=request.POST['companyname']

    c.address=request.POST['address']
    c.city=request.POST['city']
    c.state=request.POST['state']
    c.country=request.POST['country']
    c.pincode=request.POST['pincode']
    c.pan_number=request.POST['pannumber']
    c.gst_type=request.POST['gsttype']
    c.gst_no=request.POST['gstno']

    code=get_random_string(length=6)
    if company.objects.filter(Company_code = code).exists():
       code2=get_random_string(length=6)
       c.Company_code=code2
    else:
      c.Company_code=code
   
    c.save()

    staff = staff_details.objects.get(email=email,position='company',company=c)
    staff.first_name = request.POST['companyname']
    staff.last_name = ''
    staff.save()

    return redirect('login')  
  return render(request,'registercompany.html') 


def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['email']
    un=request.POST['username']
    ph=request.POST['phoneno']
    pas=request.POST['pass']
    code=request.POST['companycode']

    if company.objects.filter(Company_code=code).exists():
      com=company.objects.get(Company_code=code)
    else:
        messages.info(request, 'Sorry, Company code is Invalide')
        return redirect('registerstaff')
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un).exists():
      messages.info(request, 'Sorry, Username already exists')
      return redirect('registerstaff')
    elif staff_details.objects.filter(email=email).exists():
      messages.info(request, 'Sorry, Email already exists')
      return redirect('registerstaff')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,contact=ph,password=pas,img=img,company=com)
      staff.save()
      return redirect('login')

  else:
    print(" error")
    return redirect('registerstaff')
  

def loginurl(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['pass']
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)
        
    if staff_details.objects.filter(user_name=user_name,password=passw,position='company').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='company') 

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
        return redirect('homepage')  

    if staff_details.objects.filter(user_name=user_name,password=passw,position='staff').exists():
      data = staff_details.objects.get(user_name=user_name,password=passw,position='staff')   

      request.session["staff_id"]=data.id
      if 'staff_id' in request.session:
        if request.session.has_key('staff_id'):
          staff_id = request.session['staff_id']
          print(staff_id)
 
          return redirect('staffhome')  
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('login')  
  else:  
   return redirect('login')  
  

@login_required(login_url='login')  
def profile(request):
    if request.user.is_authenticated:
        try:
            com = company.objects.get(user=request.user)
            context = {'company': com}
            return render(request, 'profile.html', context)
        except company.DoesNotExist:
            messages.error(request, 'Company not found for the authenticated user.')
            return redirect('login')
    else:
        messages.info(request, 'Please log in to view your profile.')
        return redirect('login')
    

def editprofile(request,pk):
  com= company.objects.get(id = pk)
  context={
     'company':com
  }
  return render(request, 'editprofile.html',context)

def edit_profilesave(request,pk):
  com= company.objects.get(id = pk)
  user1 = User.objects.get(id = com.user_id)

  if request.method == "POST":

      user1.first_name = capfirst(request.POST.get('f_name'))
      user1.last_name  = capfirst(request.POST.get('l_name'))
      user1.email = request.POST.get('email')
      com.contact = request.POST.get('cnum')
      com.address = capfirst(request.POST.get('ards'))
      com.company_name = request.POST.get('comp_name')
      user1.email = request.POST.get('comp_email')
      com.city = request.POST.get('city')
      com.state = request.POST.get('state')
      com.country = request.POST.get('country')
      com.pincode = request.POST.get('pinc')
      com.gst_type = request.POST.get('gsttype')
      com.gst_no = request.POST.get('gstno')
      com.pan_number = request.POST.get('pan')
      if len(request.FILES)!=0 :
          com.profile_pic = request.FILES.get('file')

      com.save()
      user1.save()
      return redirect('profile')

  context = {
      'company' : com,
      'user1' : user1,
  } 

  return render(request,'editprofile.html',context)

def base(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context = {
              'staff' : staff

          }
  return render(request, 'base.html',context)

def staffprofile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context = {
              'staff' : staff

          }
  return render(request,'profilestaff.html',context)


def editstaffprofile(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  context={
     'staff':staff
  }
  return render(request, 'editstaff.html',context)

def edit_staffprofilesave(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)

  if request.method == "POST":

      staff.first_name = capfirst(request.POST.get('f_name'))
      staff.last_name  = capfirst(request.POST.get('l_name'))
      staff.email = request.POST.get('email')
      staff.contact = request.POST.get('cnum')
      if len(request.FILES)!=0 :
          staff.img = request.FILES.get('file')

      staff.save()
      return redirect('staffprofile')

  context = {
      'staff' : staff
  } 

  return render(request,'editstaff.html',context)


def add_debitnote(request):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  print("hii")
  print(staff)
  cmp = company.objects.get(id=staff.company.id)
  Party=party.objects.filter(company=cmp,user=cmp.user)
  allmodules = modules_list.objects.filter(company=staff.company, status='New').first()
  item=ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company)
  bank=BankModel.objects.filter(company=cmp,user=cmp.user)
  debt_count = purchasedebit.objects.filter(company=cmp).order_by('-pdebitid').first()
  
  if debt_count:
    next_count = int(debt_count.reference_number) + 1
  else:
    next_count=1

  return render(request,'adddebitnotes.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'item':item,'count':next_count,'tod':tod,'item_units':item_units,'bank':bank,'cmp':cmp})

def add_parties(request):
  
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  return render(request, 'add_parties.html',{'staff':staff})

def save_parties(request):
    if request.method == 'POST':
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstno']
        contact = request.POST['contact']
        gst_type = request.POST['gst']
        state = request.POST['state']
        address = request.POST['address']
        email = request.POST['email']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('creditlimit', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']
       
        if (
          not party_name
          
      ):
          return render(request, 'add_parties.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=staff.company.user,company=staff.company)
        part.save() 

        if 'save_and_new' in request.POST:
            
            return render(request, 'add_parties.html')
        else:
          
            return redirect('view_parties')

    return render(request, 'add_parties.html')

def create_debitnotes(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  partys=party.objects.get(id=request.POST.get('customername'))
  bill_numbers = purchasedebit.objects.filter(party=partys).values_list('billno', flat=True)

  context = {
        'bill_numbers': bill_numbers
        # Add other context variables as needed
    }
  if request.method == 'POST': 
    
    pdebt = purchasedebit(party=partys,
                      pdebitid=request.POST.get('pdebitid'),
                      debitdate=request.POST.get('debitdate'),
                      supply=request.POST.get('placosupply'),
                      payment_type=request.POST.get("method"),
                      cheque_no=request.POST.get("cheque_id"),
                      upi_no=request.POST.get("upi_id"),
                      billno=request.POST.get("bill_no"),
                      billdate=request.POST.get("billdate"), 
                      reference_number=request.POST.get("pdebitid"),
                      paid_amount = request.POST.get("advance"),
                      balance_amount = request.POST.get("balance"),
                      subtotal=float(request.POST.get('subtotal')),
                      igst = request.POST.get('igst'),
                      cgst = request.POST.get('cgst'),
                      sgst = request.POST.get('sgst'),
                      adjustment = request.POST.get("adj"),
                      taxamount = request.POST.get("taxamount"),
                      grandtotal=request.POST.get('grandtotal'),
                      company=cmp,staff=staff)
    pdebt.save()

    print(pdebt)
          
    product = tuple(request.POST.getlist("product[]"))
    qty =  tuple(request.POST.getlist("qty[]"))
    discount =  tuple(request.POST.getlist("discount[]"))
    total =  tuple(request.POST.getlist("total[]"))
    pdebitid = purchasedebit.objects.get(pdebitid=pdebt.pdebitid, company=cmp)
    



    if len(product)==len(qty)==len(discount)==len(total):
        mapped=zip(product,qty,discount,total)
        mapped=list(mapped)
        for ele in mapped:
          itm = ItemModel.objects.get(id=ele[0])
          purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],pdebit=pdebitid,company=cmp)

    purchasedebit.objects.filter(company=cmp).update(tot_debt_no=F('tot_debt_no') + 1)
          
    pdebt.tot_debt_no = pdebt.pdebitid
    pdebt.save()

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=staff,company=cmp,action='Created')

    if 'Next' in request.POST:
      return redirect('add_debitnote')
    
    if "Save" in request.POST:
      return redirect('view_purchasedebit')
    
  else:
    return render(request,'adddebitnotes.html',context)



def view_parties(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  
 
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'view_parties.html',{'staff':staff,'allmodules':allmodules,'Party':Party})

def view_party(request,id):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(company=staff.company.id)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  return render(request, 'view_party.html',{'staff':staff,'allmodules':allmodules,'Party':Party,'getparty':getparty})

def saveitem(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  name = request.POST['name']
  unit = request.POST['unit']
  hsn = request.POST['hsn']
  taxref = request.POST['taxref']
  sell_price = request.POST['sell_price']
  cost_price = request.POST['cost_price']
  intra_st = request.POST['intra_st']
  inter_st = request.POST['inter_st']

  if taxref != 'Taxable':
    intra_st = 'GST0[0%]'
    inter_st = 'IGST0[0%]'

  itmdate = request.POST.get('itmdate')
  stock = request.POST.get('stock')
  itmprice = request.POST.get('itmprice')
  minstock = request.POST.get('minstock')

  if not hsn:
    hsn = None

  itm = ItemModel(item_name=name,item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
  itm.save() 
  return JsonResponse({'success': True})




def itemdetail(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})

def savecustomer1(request):
    if request.method == 'POST':
        sid = request.session.get('staff_id')
        staff = staff_details.objects.get(id=sid)
        cmp = company.objects.get(id=staff.company.id)

        party_name = request.POST['name']
        email = request.POST['email']
        contact = request.POST['contact']
        state = request.POST['splystate']
        address = request.POST['baddress']
        gst_type = request.POST['gsttype']
        gst_no = request.POST['gstin']
        current_date = request.POST['partydate']
        openingbalance = request.POST.get('openbalance')
        payment = request.POST.get('paytype')
        creditlimit = request.POST.get('credit_limit')
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['add1']
        additionalfield2 = request.POST['add2']
        additionalfield3 = request.POST['add3']

        

        # Check if contact already exists
        if party.objects.filter(contact=contact,company=cmp).exists():
            messages.info(request, 'Sorry, Contact Number already exists')
            return redirect('add_debitnote')
        
        elif party.objects.filter(email=email, company=cmp).exists():
            messages.info(request, 'Sorry, Email already exists')
            return redirect('add_debitnote')
        else:
            part = party(party_name=party_name, gst_no=gst_no, contact=contact, gst_type=gst_type, state=state, address=address,
                         email=email, openingbalance=openingbalance, payment=payment, creditlimit=creditlimit, current_date=current_date,
                         End_date=End_date, additionalfield1=additionalfield1, additionalfield2=additionalfield2,
                         additionalfield3=additionalfield3, company=cmp, user=cmp.user)

            part.save()
            return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})


  

def cust_dropdown1(request):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  part = party.objects.filter(company=cmp,user=cmp.user)

  id_list = []
  party_list = []
  for p in part:
    id_list.append(p.id)
    party_list.append(p.party_name)

  return JsonResponse({'id_list':id_list, 'party_list':party_list })

def saveitem1(request):
  if request.method == 'POST':
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)

    name = request.POST['name']
    unit = request.POST['unit']
    hsn = request.POST['hsn']
    taxref = request.POST['taxref']
    sell_price = request.POST['sell_price']
    cost_price = request.POST['cost_price']
    intra_st = request.POST['intra_st']
    inter_st = request.POST['inter_st']

    if taxref != 'Taxable':
        intra_st = 'GST0[0%]'
        inter_st = 'IGST0[0%]'

    itmdate = request.POST.get('itmdate')
    stock = request.POST.get('stock')
    itmprice = request.POST.get('itmprice')
    minstock = request.POST.get('minstock')

    # Check if the HSN already exists
    if ItemModel.objects.filter(item_hsn=hsn,company=cmp).exists():
       messages.info(request, 'Sorry, HSN Number already exists')
       return redirect('add_debitnote')
    else:
        itm = ItemModel(item_name=name,item_hsn=hsn,item_unit=unit,item_taxable=taxref, item_gst=intra_st,item_igst=inter_st, item_sale_price=sell_price, 
                    item_purchase_price=cost_price,item_opening_stock=stock,item_current_stock=stock,item_at_price=itmprice,item_date=itmdate,
                    item_min_stock_maintain=minstock,company=cmp,user=cmp.user)
        itm.save() 
        return JsonResponse({'success': True})
    
  else:
        return JsonResponse({'success': False, 'message': 'Invalid request method'})


def item_dropdowns(request):
  if request.user.is_company:
      cmp = request.user.company
  else:
      cmp = request.user.employee.company
  options = {}
  option_objects = ItemModel.objects.all(company=cmp)
  for option in option_objects:
      options[option.id] = [option.item_name]
  return JsonResponse(options)

def custdata1(request):
  cid = request.POST['id']
  part = party.objects.get(id=cid)
  # email = part.email
  phno = part.contact
  address = part.address
  pay = part.payment
  bal = part.openingbalance
  return JsonResponse({ 'phno':phno, 'address':address, 'pay':pay, 'bal':bal})

def purchasebilldata(request):
    try:
        party_name = request.POST['id']
        party_instance = party.objects.get(id=party_name)

        # Initialize lists to store multiple bill numbers and dates
        bill_numbers = []
        bill_dates = []

        try:
            # Retrieve all PurchaseBill instances for the party
            bill_instances = PurchaseBill.objects.filter(party=party_instance)

            # Loop through each PurchaseBill instance and collect bill numbers and dates
            for bill_instance in bill_instances:
                bill_numbers.append(bill_instance.billno)
                bill_dates.append(bill_instance.billdate)

        except PurchaseBill.DoesNotExist:
            pass

        # Return a JSON response with the list of bill numbers and dates
        if not bill_numbers and not bill_dates:
            return JsonResponse({'bill_numbers': ['nobill'], 'bill_dates': ['nodate']})

        return JsonResponse({'bill_numbers': bill_numbers, 'bill_dates': bill_dates})

    except KeyError:
        return JsonResponse({'error': 'The key "id" is missing in the POST request.'})

    except party.DoesNotExist:
        return JsonResponse({'error': 'Party not found.'})
    
def get_bill_date(request):
    selected_bill_no = request.POST.get('bill_no', None)

    try:
        # Get the latest PurchaseBill with the specified bill_number
        purchase_bill = PurchaseBill.objects.filter(billno=selected_bill_no).latest('billdate')
        bill_date = purchase_bill.billdate.strftime('%Y-%m-%d')
    except PurchaseBill.DoesNotExist:
        return JsonResponse({'error': 'Bill number not found'}, status=400)
    except PurchaseBill.MultipleObjectsReturned:
        # Handle the case where multiple PurchaseBills are found for the same bill_number
        return JsonResponse({'error': 'Multiple PurchaseBills found for the same bill number'}, status=400)

    return JsonResponse({'bill_date': bill_date})


def bankdata1(request):
  bid = request.POST['id']
  bank = BankModel.objects.get(id=bid) 
  bank_no = bank.account_num
  return JsonResponse({'bank_no':bank_no})



def itemdetail(request):
  itmid = request.GET['id']
  itm = ItemModel.objects.get(id=itmid)
  hsn = itm.item_hsn
  gst = itm.item_gst
  igst = itm.item_igst
  price = itm.item_purchase_price
  qty = itm.item_current_stock
  return JsonResponse({'hsn':hsn, 'gst':gst, 'igst':igst, 'price':price, 'qty':qty})


def debthistory(request):
    if request.method == 'POST' and 'id' in request.POST:
        pid = request.POST['id']
        sid = request.session.get('staff_id')
        
        if sid:
            staff = get_object_or_404(staff_details, id=sid)
            cmp = get_object_or_404(company, id=staff.company.id)
            pdebt = get_object_or_404(purchasedebit, pdebitid=pid, company=cmp)
            hsty = DebitnoteTransactionHistory.objects.filter(debitnote=pdebt, company=cmp).last()
            
            if hsty:
                name = hsty.staff.first_name + ' ' + hsty.staff.last_name 
                action = hsty.action
                return JsonResponse({'name': name, 'action': action, 'pid': pid})
            else:
                return JsonResponse({'error': 'No transaction history found for this purchase debit.'}, status=404)
        else:
            return JsonResponse({'error': 'Staff ID not found in session.'}, status=404)
    else:
        return JsonResponse({'error': 'Invalid request method or missing ID parameter.'}, status=400)

def delete_debit(request,id):
    sid = request.session.get('staff_id')
    staff = get_object_or_404(staff_details, id=sid)
    cmp = get_object_or_404(company, id=staff.company.id)

    # Retrieve the purchasedebit instance to be deleted
    pdebt = get_object_or_404(purchasedebit, pdebitid=id)

    # Ensure the purchasedebit instance belongs to the user's company
    if pdebt.company != cmp:
        return redirect('view_purchasedebit')  # Redirect to view_purchasedebit if not authorized

    # Delete related purchasedebit1 instances and the purchasedebit instance itself
    purchasedebit1.objects.filter(pdebit=pdebt, company=cmp).delete()
    pdebt.delete()

    return redirect('view_purchasedebit')


def details_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id) 
  allmodules = modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
  dis = 0
  for itm in pitm:
    dis += int(itm.discount)
  itm_len = len(pitm)

  context={'staff':staff,'allmodules':allmodules,'pdebt':pdebt,'pitm':pitm,'itm_len':itm_len,'dis':dis}
  return render(request,'debitnotedetails.html',context)

def edit_debitnote(request,id):
  toda = date.today()
  tod = toda.strftime("%Y-%m-%d")
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)
  partys = party.objects.filter(company=cmp,user=cmp.user)
  item = ItemModel.objects.filter(company=cmp,user=cmp.user)
  item_units = UnitModel.objects.filter(user=cmp.user,company=staff.company.id)
  allmodules= modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  debtitem = purchasedebit1.objects.filter(pdebit=id,company=cmp)
  ddate = pdebt.debitdate.strftime("%Y-%m-%d")



  context = {'staff':staff,  'allmodules':allmodules, 'pdebt':pdebt, 'debtitem':debtitem, 'partys':partys, 'item':item, 'item_units':item_units, 'ddate':ddate,'tod':tod}
  return render(request,'debitnoteedit.html',context)


def history_debitnote(request,id):
  sid = request.session.get('staff_id')
  staff =  staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)  
  allmodules= modules_list.objects.filter(company=staff.company,status='New')
  pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
  hsty= DebitnoteTransactionHistory.objects.filter(debitnote=id,company=cmp)
  context = {'staff':staff,'allmodules':allmodules,'hsty':hsty,'id':id}
  return render(request,'debitnotehistory.html',context)

def update_debitnote(request,id):
  if request.method =='POST':
    sid = request.session.get('staff_id')
    staff = staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)  
    partys = party.objects.get(id=request.POST.get('customername'))
    pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
    pdebt.party = partys
    pdebt.debitdate = request.POST.get('debitdate')
    pdebt.billno = request.POST.get('bill_no')
    pdebt.billdate = request.POST.get('billdate')
    pdebt.supply  = request.POST.get('placosupply')
    pdebt.subtotal =float(request.POST.get('subtotal'))
    pdebt.grandtotal = request.POST.get('grandtotal')
    pdebt.igst = request.POST.get('igst')
    pdebt.cgst = request.POST.get('cgst')
    pdebt.sgst = request.POST.get('sgst')
    pdebt.taxamount = request.POST.get("taxamount")
    pdebt.adjustment = request.POST.get("adj")
    pdebt.payment_type = request.POST.get("method")
    pdebt.cheque_no = request.POST.get("cheque_id")
    pdebt.upi_no = request.POST.get("upi_id")
    pdebt.paid_amount = request.POST.get("advance")
    pdebt.balance_amount = request.POST.get("balance")

    pdebt.save()

    product = tuple(request.POST.getlist("product[]"))
    qty = tuple(request.POST.getlist("qty[]"))
    total = tuple(request.POST.getlist("total[]"))
    discount = tuple(request.POST.getlist("discount[]"))

    purchasedebit1.objects.filter(pdebit=pdebt,company=cmp).delete()
    if len(total)==len(discount)==len(qty):
      mapped=zip(product,qty,discount,total)
      mapped=list(mapped)
      for ele in mapped:
        itm = ItemModel.objects.get(id=ele[0])
        purchasedebit1.objects.create(product =itm,qty=ele[1],discount=ele[2],total=ele[3],pdebit=pdebt,company=cmp)

    DebitnoteTransactionHistory.objects.create(debitnote=pdebt,company=cmp,staff=staff,action='Updated')
    return redirect('view_purchasedebit')

  return redirect('view_purchasedebit')


def sharedebitToEmail(request,id):
  if request.user:
        try:
            if request.method == 'POST':
                emails_string = request.POST['email_ids']

                # Split the string by commas and remove any leading or trailing whitespace
                emails_list = [email.strip() for email in emails_string.split(',')]
                email_message = request.POST['email_message']
                print(emails_list)

                sid = request.session.get('staff_id')
                staff =  staff_details.objects.get(id=sid)
                cmp = company.objects.get(id=staff.company.id) 
               
                pdebt = purchasedebit.objects.get(pdebitid=id,company=cmp)
                pitm = purchasedebit1.objects.filter(pdebit=pdebt,company=cmp)
                        
                context = {'pdebt':pdebt, 'cmp':cmp,'pitm':pitm}
                template_path ='debitnote_file_mail.html'
                template = get_template(template_path)

                html  = template.render(context)
                result = BytesIO()
                pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
                pdf = result.getvalue()
                filename = f'DEBIT NOTE - {pdebt.pdebitid}.pdf'
                subject = f"DEBIT NOTE - {pdebt.pdebitid}"
                email = EmailMessage(subject, f"Hi,\nPlease find the attached DEBIT NOTE - File-{pdebt.pdebitid}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
                email.attach(filename, pdf, "application/pdf")
                email.send(fail_silently=False)

                msg = messages.success(request, 'Debit note file has been shared via email successfully..!')
                return redirect(details_debitnote,id)
        except Exception as e:
            print(e)
            messages.error(request, f'{e}')
            return redirect(details_debitnote, id)
        

def view_purchasedebit(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)
  cmp = company.objects.get(id=staff.company.id)
  print("hello")
  print(staff)
  allmodules = modules_list.objects.filter(company=staff.company, status='New').first()
  pdebt = purchasedebit.objects.filter(company=cmp)

  if not pdebt:
    context = {'staff':staff, 'allmodules':allmodules}
    return render(request,'emptydebit.html',context)
  
  context = {'staff':staff,'allmodules':allmodules,'pdebt':pdebt}
  return render(request,'purchase_return_dr.html',context)


def viewPaymentIn(request,id):
  sid = request.session.get('staff_id')
  staff = staff_details.objects.get(id=sid)
  cmp = company.objects.get(id=staff.company.id)

  paymentInDetails = PaymentIn.objects.get(id = id)
  allmodules= modules_list.objects.get(company=cmp,status='New')
  context = {
    'payment':paymentInDetails,'staff':staff,'allmodules':allmodules,'company':cmp,
  }

  return render(request, 'payment_in_details.html',context)

def sharePaymentInToEmail(request,id):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
    else:
      return redirect('/')
    staff =  staff_details.objects.get(id=staff_id)
    com =  company.objects.get(id = staff.company.id)
    try:
      if request.method == 'POST':
        emails_string = request.POST['email_ids']

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(',')]
        email_message = request.POST['email_message']
        # print(emails_list)

        payment = PaymentIn.objects.get(id = id)
        context = {'payment': payment,'company':com}
        template_path = 'payment_in_pdf.html'
        template = get_template(template_path)

        html  = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f'Payment In - {payment.rec_no}.pdf'
        subject = f"Payment In Receipt - {payment.rec_no}"
        email = EmailMessage(subject, f"Hi,\nPlease find the attached Receipt of Payment In -{payment.rec_no}. \n{email_message}\n\n--\nRegards,\n{com.company_name}\n{com.address}\n{com.city} - {com.state}\n{com.contact}", from_email=settings.EMAIL_HOST_USER, to=emails_list)
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        messages.success(request, 'Receipt has been shared via email successfully..!')
        return redirect(viewPaymentIn,id)
    except Exception as e:
        print(e)
        messages.error(request, f'{e}')
        return redirect(viewPaymentIn, id)
    


def import_debitnote(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(purchasedebit.objects.filter(company=cmp).last().tot_debt_no)

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      debitsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=debitsheet[0],email=debitsheet[1],company=cmp)
      purchasedebit.objects.create(party=part,pdebitid = totval,
                                  debitdate=debitsheet[2],
                                  supply =debitsheet[3],
                                  tot_debt_no = totval,
                                  company=cmp,staff=staff)
      
      pdebt = purchasedebit.objects.last()
      if debitsheet[4] == 'Cheque':
        pdebt.payment_type = 'Cheque'
        pdebt.cheque_no = debitsheet[5]
      elif debitsheet[4] == 'UPI':
        pdebt.upi_no = debitsheet[5]
      else:
        if debitsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=debitsheet[4],company=cmp)
          pdebt.payment_type = bank
        else:
          pdebt.payment_type = 'Cash'
      pdebt.save()

      purchasedebit.objects.filter(company=cmp).update(tot_debt_no=totval )
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],hsn=prdsheet[2],company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          
          purchasedebit1.objects.create(pdebit=pdebt,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

       
          if debitsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if debitsheet[3]=='State':
            gst = round((taxamount/2),2)
            pdebt.sgst=gst
            pdebt.cgst=gst
            pdebt.igst=0

          else:
            gst=round(taxamount,2)
            pdebt.igst=gst
            pdebt.cgst=0
            pdebt.sgst=0

      gtotal = subtotal + taxamount + float(debitsheet[6])
      balance = gtotal- float(debitsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pdebt.subtotal=round(subtotal,2)
      pdebt.taxamount=round(taxamount,2)
      pdebt.adjustment=round(debitsheet[6],2)
      pdebt.grandtotal=gtotal
      pdebt.paid_amount=round(debitsheet[7],2)
      pdebt.balance_amount=balance
      pdebt.save()

      DebitnoteTransactionHistory.objects.create(debitnote=pdebt,staff=pdebt.staff,company=pdebt.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
   

def import_purchase_bill(request):
  if request.method == 'POST' and request.FILES['billfile']  and request.FILES['prdfile']:
    sid = request.session.get('staff_id')
    staff =  staff_details.objects.get(id=sid)
    cmp = company.objects.get(id=staff.company.id)
    totval = int(PurchaseBill.objects.filter(company=cmp).last().tot_bill_no) + 1

    excel_bill = request.FILES['billfile']
    excel_b = load_workbook(excel_bill)
    eb = excel_b['Sheet1']
    excel_prd = request.FILES['prdfile']
    excel_p = load_workbook(excel_prd)
    ep = excel_p['Sheet1']

    for row_number1 in range(2, eb.max_row + 1):
      billsheet = [eb.cell(row=row_number1, column=col_num).value for col_num in range(1, eb.max_column + 1)]
      part = party.objects.get(party_name=billsheet[0],email=billsheet[1],company=cmp)
      PurchaseBill.objects.create(party=part,billno=totval,
                                  billdate=billsheet[2],
                                  supplyplace =billsheet[3],
                                  tot_bill_no = totval,
                                  company=cmp,staff=staff)
      
      pbill = PurchaseBill.objects.last()
      if billsheet[4] == 'Cheque':
        pbill.pay_method = 'Cheque'
        pbill.cheque_no = billsheet[5]
      elif billsheet[4] == 'UPI':
        pbill.pay_method = 'UPI'
        pbill.upi_no = billsheet[5]
      else:
        if billsheet[4] != 'Cash':
          bank = BankModel.objects.get(bank_name=billsheet[4],company=cmp)
          pbill.pay_method = bank
        else:
          pbill.pay_method = 'Cash'
      pbill.save()

      PurchaseBill.objects.filter(company=cmp).update(tot_bill_no=totval)
      totval += 1
      subtotal = 0
      taxamount=0
      for row_number2 in range(2, ep.max_row + 1):
        prdsheet = [ep.cell(row=row_number2, column=col_num).value for col_num in range(1, ep.max_column + 1)]
        if prdsheet[0] == row_number1:
          itm = ItemModel.objects.get(item_name=prdsheet[1],hsn=int(prdsheet[2]),company=cmp)
          total=int(prdsheet[3])*int(itm.item_purchase_price) - int(prdsheet[4])
          PurchaseBillItem.objects.create(purchasebill=pbill,
                                company=cmp,
                                product=itm,
                                qty=prdsheet[3],
                                discount=prdsheet[4],
                                total=total)

          if billsheet[3] =='State':
            taxval = itm.item_gst
            taxval=taxval.split('[')
            tax=int(taxval[0][3:])
          else:
            taxval = itm.item_igst
            taxval=taxval.split('[')
            tax=int(taxval[0][4:])

          subtotal += total
          tamount = total *(tax / 100)
          taxamount += tamount
                
          if billsheet[3]=='State':
            gst = round((taxamount/2),2)
            pbill.sgst=gst
            pbill.cgst=gst
            pbill.igst=0

          else:
            gst=round(taxamount,2)
            pbill.igst=gst
            pbill.cgst=0
            pbill.sgst=0

      gtotal = subtotal + taxamount + float(billsheet[6])
      balance = gtotal- float(billsheet[7])
      gtotal = round(gtotal,2)
      balance = round(balance,2)

      pbill.subtotal=round(subtotal,2)
      pbill.taxamount=round(taxamount,2)
      pbill.adjust=round(billsheet[6],2)
      pbill.grandtotal=gtotal
      pbill.advance=round(billsheet[7],2)
      pbill.balance=balance
      pbill.save()

      PurchaseBillTransactionHistory.objects.create(purchasebill=pbill,staff=pbill.staff,company=pbill.company,action='Created')
      return JsonResponse({'message': 'File uploaded successfully!'})
  else:
    return JsonResponse({'message': 'File upload Failed!'})
  
def check_contact_exists(request):
    contact = request.GET.get('contact')
    if party.objects.filter(contact=contact).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})

def check_email_exists(request):
    email= request.GET.get('email')
    if party.objects.filter(email=email).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})
  

def check_hsn_exists(request):
    hsn= request.GET.get('hsn')
    
    if ItemModel.objects.filter(item_hsn=hsn).exists():
        return JsonResponse({'exists': True})
    return JsonResponse({'exists':False})


  


    